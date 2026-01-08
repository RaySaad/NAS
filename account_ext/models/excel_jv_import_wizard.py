# -*- coding: utf-8 -*-

import base64
import json
import logging
from datetime import datetime
from odoo import api, models, fields, _
from odoo.exceptions import ValidationError
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)


class ExcelJVImportWizard(models.TransientModel):
    _name = 'excel.jv.import.wizard'
    _description = 'Excel JV Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    jv_id = fields.Many2one('account.move', string='Created Journal Entry', readonly=True)

    def action_import_excel(self):
        """
        Process Excel file and create Journal Entry dynamically.
        
        Structure:
        - First row: Headers
        - First 3 columns: Journal Entry level fields (date, reference, journal)
        - Remaining columns: Journal Item level fields
        - Special handling: Account Code and Operating Unit Code lookup
        """
        if not openpyxl:
            raise ValidationError(_('Please install openpyxl library: pip install openpyxl'))

        if not self.excel_file:
            raise ValidationError(_('Please upload an Excel file'))

        # Decode the file
        try:
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            worksheet = workbook.active
        except Exception as e:
            raise ValidationError(_('Error reading Excel file: %s') % str(e))

        # Read header row
        headers = []
        header_row = 1
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            headers.append(str(cell_value).strip() if cell_value else '')

        if not headers or all(not h for h in headers):
            raise ValidationError(_('No headers found in Excel file. Please ensure the first row contains column headers.'))

        # Separate header-level (first 3 columns) and line-level (remaining columns)
        move_headers = headers[:3]  # First 3 columns for Journal Entry
        line_headers = headers[3:]  # Remaining columns for Journal Items

        # Get model fields for dynamic mapping
        move_fields = self.env['account.move']._fields
        line_fields = self.env['account.move.line']._fields

        # Process data rows
        move_data = {}
        line_vals_list = []
        errors = []
        first_data_row = None

        for row in range(header_row + 1, worksheet.max_row + 1):
            # Check if row is empty
            row_empty = True
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    row_empty = False
                    break
            
            if row_empty:
                continue

            # Track first data row for move-level fields
            if first_data_row is None:
                first_data_row = row

            # Extract move-level data from first 3 columns (only from first data row)
            if row == first_data_row:
                for col_idx, header in enumerate(move_headers, start=1):
                    if not header:
                        continue
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    if cell_value is not None:
                        field_name = self._normalize_field_name(header)
                        if field_name in move_fields:
                            move_data[field_name] = self._convert_field_value(cell_value, move_fields[field_name], field_name)
                        else:
                            # Try to map common field names
                            mapped_field = self._map_header_to_field(header, move_fields)
                            if mapped_field:
                                move_data[mapped_field] = self._convert_field_value(cell_value, move_fields[mapped_field], mapped_field)

            # Extract line-level data from remaining columns
            line_data = {}
            for col_idx, header in enumerate(line_headers, start=4):  # Start from column 4
                if not header:
                    continue
                cell_value = worksheet.cell(row=row, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    field_name = self._normalize_field_name(header)
                    line_data[header] = cell_value  # Keep original header for special processing

            if not line_data:
                continue

            # Process line item
            try:
                line_val = self._process_line_item(line_data, line_fields, row)
                if line_val:
                    line_vals_list.append((0, 0, line_val))
            except ValidationError as e:
                errors.append(_('Row %s: %s') % (row, str(e)))
            except Exception as e:
                _logger.exception('Error processing row %s', row)
                errors.append(_('Row %s: Unexpected error - %s') % (row, str(e)))

        if errors:
            error_msg = '\n'.join(errors)
            raise ValidationError(_('Errors found while processing Excel file:\n%s') % error_msg)

        if not line_vals_list:
            raise ValidationError(_('No valid data found in Excel file'))


        if 'move_type' not in move_data:
            move_data['move_type'] = 'entry'

        # Create Journal Entry
        move_data['line_ids'] = line_vals_list
        jv = self.env['account.move'].create(move_data)
        for line in jv.line_ids:
            if line.employee_id:
                line.customer_account = line.employee_id.customer_account.id
                line.partner_id = line.employee_id.customer_account.partner_id.id
                line.customer_code = line.employee_id.customer_account.partner_id.customer_code
            elif line.customer_account:
                line.partner_id = line.employee_id.customer_account.partner_id.id
                line.customer_code = line.employee_id.customer_account.partner_id.customer_code
        try:
            jv.ref = str(int(float(jv.ref)))
        except:
            jv.ref = str(jv.ref)
        self.jv_id = jv.id

        return {
            'type': 'ir.actions.act_window',
            'name': _('Journal Entry Created'),
            'res_model': 'account.move',
            'res_id': jv.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _parse_date(self, date_value):
        """Parse date from various formats"""
        if isinstance(date_value, datetime):
            return date_value.date()
        elif isinstance(date_value, str):
            # Try common date formats
            date_formats = [
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%m/%d/%Y',
                '%d-%m-%Y',
                '%m-%d-%Y',
                '%Y/%m/%d',
            ]
            for fmt in date_formats:
                try:
                    return datetime.strptime(date_value.strip(), fmt).date()
                except:
                    continue
        return fields.Date.today()

    def _process_line_item(self, line_data, line_fields, row_num):
        """
        Process a single line item from Excel and return line values.
        
        Special handling:
        - Account Code -> search account.account by code
        - Operating Unit Code -> search operating.unit by code
        - All other fields -> dynamic mapping based on header name
        """
        line_val = {}
        row_info = ' (Row %s)' % row_num if row_num else ''

        # Special handling for Account Code
        account_found = False
        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'account_id' == header_lower:
                account_code = str(int(value)).strip()
                if account_code:
                    account = self.env['account.account'].search([
                        ('code', '=', account_code)
                    ], limit=1)
                    if not account:
                        raise ValidationError(_('Account with code "%s" not found%s') % (account_code, row_info))
                    line_val['account_id'] = account.id
                    account_found = True
                    break

        if not account_found:
            raise ValidationError(_('Account Code is required. Please provide "Account Code" column%s') % row_info)

        # Special handling for Operating Unit Code
        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'operating_unit_id' in header_lower:
                ou_code = str(value).strip()
                if ou_code:
                    operating_unit = self.env['operating.unit'].search([
                        ('code', '=', ou_code)
                    ], limit=1)
                    if not operating_unit:
                        raise ValidationError(_('Operating Unit with code "%s" not found%s') % (ou_code, row_info))
                    line_val['operating_unit_id'] = operating_unit.id
                    break

        # Special handling for Operating Unit Code
        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'customer_account' in header_lower:
                name = str(value).strip()
                if name:
                    customer_account = self.env['partner.subscription'].search([
                        ('name', '=', name)
                    ], limit=1)
                    if not customer_account:
                        raise ValidationError(
                            _('Customer Account with name "%s" not found%s') % (name, row_info))
                    line_val['customer_account'] = customer_account.id
                    break

        # Special handling for Operating Unit Code
        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'vehicle_id' in header_lower:
                name = value
                if name:
                    vehicle_id = self.env['fleet.vehicle'].search([
                        ('display_name', '=', name)
                    ], limit=1)
                    if not vehicle_id:
                        raise ValidationError(
                            _('Vehicle with display name "%s" not found%s') % (name, row_info))
                    line_val['vehicle_id'] = vehicle_id.id
                    break
        # Special handling for Operating Unit Code
        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'employee_code' in header_lower:
                name = value
                if name:
                    employee_id = self.env['hr.employee'].search([
                        ('employee_code', '=', name)
                    ], limit=1)
                    if not employee_id:
                        raise ValidationError(
                            _('Employee having employee_code "%s" not found%s') % (name, row_info))
                    line_val['employee_id'] = employee_id.id
                    break

        for header, value in line_data.items():
            header_lower = header.lower().strip()
            if 'employee_code' in header_lower:
                name = value
                if name:
                    employee_id = self.env['hr.employee'].search([
                        ('employee_code', '=', name)
                    ], limit=1)
                    if not employee_id:
                        raise ValidationError(
                            _('Employee having employee_code "%s" not found%s') % (name, row_info))
                    line_val['employee_id'] = employee_id.id
                    break

        # Process all other fields dynamically
        for header, value in line_data.items():
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            header_lower = header.lower().strip()
            
            # Skip already processed fields
            if header_lower in ['account_id', 'operating_unit_id', 'customer_account', 'vehicle_id', 'employee_code']:
                continue

            # Normalize field name
            field_name = self._normalize_field_name(header)
            
            # Try to find matching field
            mapped_field = self._map_header_to_field(header, line_fields)
            if not mapped_field:
                # Field doesn't exist, skip it
                _logger.debug('Field "%s" not found in account.move.line, skipping', header)
                continue
            if mapped_field == 'tax_tag_ids':
                line_val[mapped_field] = [(6, 0, [int(value)])]
            elif mapped_field == 'analytic_distribution':
                line_val[mapped_field] = json.loads(value)
            else:
                line_val[mapped_field] = value

        return line_val

    def _normalize_field_name(self, header):
        """Convert header to potential field name"""
        # Remove special characters, convert to lowercase, replace spaces with underscores
        normalized = header.lower().strip()
        normalized = normalized.replace(' ', '_').replace('-', '_')
        # Remove common prefixes/suffixes
        normalized = normalized.replace('journal_items/', '').replace('journal_item/', '')
        return normalized

    def _map_header_to_field(self, header, fields_dict):
        """Map header name to actual field name"""
        header_lower = header.lower().strip()
        normalized = self._normalize_field_name(header)

        # Direct match
        if normalized in fields_dict:
            return normalized

        # Common mappings
        mappings = {
            'name': ['name', 'description', 'label', 'narration'],
            'debit': ['debit', 'dr'],
            'credit': ['credit', 'cr'],
            'partner_id': ['partner', 'customer', 'vendor', 'supplier'],
            'tax_ids': ['tax', 'taxes', 'tax_id'],
            'analytic_distribution': ['analytic_distribution', 'analytic', 'analytic_account'],
            'employee_id': ['employee', 'employee_id'],
            'vehicle_id': ['vehicle', 'fleet_vehicle'],
        }

        for field_name, aliases in mappings.items():
            if field_name in fields_dict:
                for alias in aliases:
                    if alias in header_lower or alias in normalized:
                        return field_name

        return None

    def _convert_field_value(self, value, field, field_name, row_info=''):
        """Convert Excel value to appropriate field type"""
        if value is None:
            return None

        # Handle Many2one fields
        if field.type == 'many2one':
            # Try to find by name or code
            comodel = self.env[field.comodel_name]
            # Check if model has 'code' field
            if 'code' in comodel._fields:
                record = comodel.search([('code', '=', str(value).strip())], limit=1)
                if record:
                    return record.id
            # Try by name
            record = comodel.search([('name', '=', str(value).strip())], limit=1)
            if record:
                return record.id
            # If not found and field is not required, return False
            if not field.required:
                return False
            raise ValidationError(_('Record not found for %s: %s%s') % (field_name, value, row_info))

        # Handle Many2many fields
        if field.type == 'many2many':
            # Expect comma-separated values or list
            if isinstance(value, str):
                values = [v.strip() for v in value.split(',')]
            else:
                values = [value]
            
            record_ids = []
            comodel = self.env[field.comodel_name]
            for val in values:
                if 'code' in comodel._fields:
                    record = comodel.search([('code', '=', str(val).strip())], limit=1)
                    if record:
                        record_ids.append(record.id)
                        continue
                record = comodel.search([('name', '=', str(val).strip())], limit=1)
                if record:
                    record_ids.append(record.id)
            
            if record_ids:
                return [(6, 0, record_ids)]
            return [(5,)]  # Clear all

        # Handle One2many fields (usually not set directly)
        if field.type == 'one2many':
            return None

        # Handle Boolean
        if field.type == 'boolean':
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.lower() in ('true', '1', 'yes', 'y')
            return bool(value)

        # Handle Integer
        if field.type == 'integer':
            try:
                return int(float(value))
            except:
                return 0

        # Handle Float/Monetary
        if field.type in ('float', 'monetary'):
            try:
                return float(value)
            except:
                return 0.0

        # Handle Date
        if field.type == 'date':
            return self._parse_date(value)

        # Handle Datetime
        if field.type == 'datetime':
            if isinstance(value, datetime):
                return value
            # Try to parse
            try:
                return datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S')
            except:
                return self._parse_date(value)

        # Handle Selection
        if field.type == 'selection':
            # Return as-is, validation will happen at create
            return str(value).strip()

        # Handle JSON (analytic_distribution)
        if field_name == 'analytic_distribution' and isinstance(value, str):
            try:
                import json
                return json.loads(value)
            except:
                # Try to find analytic account by code
                analytic_account = self.env['account.analytic.account'].search([
                    ('code', '=', str(value).strip())
                ], limit=1)
                if analytic_account:
                    return {str(analytic_account.id): 100.0}
                return {}

        # Default: return as string
        return str(value).strip()

    def _find_journal(self, journal_identifier):
        """Find journal by code or name"""
        # Try by code first
        journal = self.env['account.journal'].search([
            ('code', '=', journal_identifier)
        ], limit=1)
        if journal:
            return journal
        
        # Try by name
        journal = self.env['account.journal'].search([
            ('name', '=', journal_identifier)
        ], limit=1)
        if journal:
            return journal
        
        # Try case-insensitive
        journal = self.env['account.journal'].search([
            ('name', 'ilike', journal_identifier)
        ], limit=1)
        return journal

    def _convert_value(self, value, field):
        """Legacy method for backward compatibility"""
        return self._convert_field_value(value, field, field.name, '')

